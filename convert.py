"""
Shree Narayani Traders — Data Converter
========================================
Run this any time you:
  • Update Master_File.xlsx (add/edit/remove products, including MRP values)
  • Add a new company to the spreadsheet
  • Drop new photos into the Photos/ folder

It regenerates data.js which the website reads.

Usage:
    python convert.py
    python convert.py --file "C:/path/to/Master_File.xlsx"

Excel column order (Master_File.xlsx):
    Company | PRODUCTS | PACK | Composition | In Boxes | Shipper Qty | PTS | MRP | Remark

Photo naming:
    Save product photos in a "Photos/" folder next to this script.
    Name each file exactly after the product name in Excel.
    All of these work for a product called "AMOXYCLAV 625":
        Photos/AMOXYCLAV 625.jpg
        Photos/AMOXYCLAV 625.png
        Photos/AMOXYCLAV 625.webp
    Supported extensions: jpg, jpeg, png, webp, gif
"""

import json, sys, os, re, argparse
from datetime import datetime
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl


# ── Sanitize function ────────────────────────────────────────────────────────
# IMPORTANT: This must produce identical output to sanitizePhotoName() in
# SNT.html so that PHOTO_MAP keys match the JS lookups.
def sanitize_photo_name(name: str) -> str:
    """Strip Windows/Linux illegal filename chars, trim whitespace, lowercase."""
    return re.sub(r'[/\\:*?"<>|]', '', name).strip().lower()


def parse_float(v):
    """Best-effort float parse. Returns None for blank / unparseable cells."""
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(',', '').replace('₹', '')
    if not s: return None
    try:
        return float(s)
    except ValueError:
        return None


# ── Main convert ─────────────────────────────────────────────────────────────
def convert(excel_path="Master_File.xlsx", output_path="data.js"):

    if not os.path.exists(excel_path):
        print(f"ERROR: File not found → {excel_path}")
        sys.exit(1)

    print(f"\nReading: {excel_path}")
    wb   = openpyxl.load_workbook(excel_path, data_only=True)
    data = []
    sr   = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Column order: Company | Product | Pack | Composition | InBoxes | ShipperQty | PTS | MRP | Remark
            if row[0] is None or str(row[0]).strip() == "":
                continue
            # Skip repeated header rows
            if str(row[0]).strip().lower() in ("company", "sr.no.", "sr no"):
                continue
            sr += 1
            pts = parse_float(row[6])
            # Tolerate older files without MRP column (length 8)
            mrp = parse_float(row[7]) if len(row) >= 8 else None
            remark = row[8] if len(row) >= 9 else (row[7] if len(row) >= 8 and not isinstance(row[7], (int, float)) else None)
            # Heuristic: if row has 8 cols and row[7] is text (not numeric), treat as remark (backward compat)
            if len(row) == 8:
                # Old format — column 8 is remark
                mrp = None
                remark = row[7]

            data.append({
                "sr":          sr,
                "company":     str(row[0]).strip(),
                "product":     str(row[1]).strip()                    if row[1] else "",
                "pack":        str(row[2]).strip()                    if row[2] else "",
                "composition": str(row[3]).replace("\n", " ").strip() if row[3] else "",
                "inBoxes":     str(row[4])   if row[4] else None,
                "shipperQty":  str(row[5])   if row[5] else None,
                "pts":         pts,
                "mrp":         mrp,
                "remark":      str(remark).strip() if remark else "",
            })

    updated_at = datetime.now().strftime("%d %b %Y, %I:%M %p")
    # Companies list — sorted; order defines which color each company gets in the UI
    companies  = sorted(set(d["company"] for d in data if d["company"]))

    # ── Scan Photos/ folder ──────────────────────────────────────────────────
    photo_exts  = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    photos_dir  = os.path.join(script_dir, "Photos")

    # photo_map: sanitized_product_name → actual_filename_with_extension
    # SNT.html looks up sanitizePhotoName(product) → "Photos/" + photo_map[key]
    photo_map = {}
    if os.path.isdir(photos_dir):
        for fname in os.listdir(photos_dir):
            name_part, ext = os.path.splitext(fname)
            if ext.lower() in photo_exts and name_part.strip():
                key = sanitize_photo_name(name_part)
                if key:
                    photo_map[key] = fname  # last-write wins for same key

    # ── Write data.js ────────────────────────────────────────────────────────
    js = f"""// Auto-generated by convert.py — DO NOT EDIT MANUALLY
// Last updated: {updated_at}

const CATALOGUE_DATA = {json.dumps(data, indent=2, ensure_ascii=False)};

const CATALOGUE_META = {{
  updatedAt: "{updated_at}",
  totalProducts: {len(data)},
  companies: {json.dumps(companies, ensure_ascii=False)}
}};

// Maps sanitized product name → actual photo filename in Photos/
// Lookup: PHOTO_MAP[sanitizePhotoName(product)] → filename
const PHOTO_MAP = {json.dumps(photo_map, indent=2, ensure_ascii=False)};
"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(js)

    # ── Console report ───────────────────────────────────────────────────────
    by_co = Counter(d["company"] for d in data)
    print(f"\n✓ {len(data)} products · {len(companies)} companies\n")
    for i, c in enumerate(companies):
        print(f"    [{i+1}] {c}: {by_co[c]} products")

    print(f"\n✓ Output  : {output_path}")
    print(f"✓ Updated : {updated_at}")

    # ── MRP coverage report ──────────────────────────────────────────────────
    with_mrp    = [d for d in data if d["mrp"] is not None]
    without_mrp = [d for d in data if d["mrp"] is None]
    print(f"\n💰 MRP coverage: {len(with_mrp)} / {len(data)} products")
    if without_mrp and len(without_mrp) <= 15:
        print(f"   Missing MRP for {len(without_mrp)} products:")
        for d in without_mrp:
            print(f"      → {d['product']}")
    elif without_mrp:
        print(f"   {len(without_mrp)} products still need MRP — fill the MRP column in Excel and rerun.")

    # Average margin where both PTS and MRP exist
    margins = []
    for d in data:
        if d['pts'] and d['mrp'] and d['pts'] > 0:
            margins.append((d['mrp'] - d['pts']) / d['pts'] * 100)
    if margins:
        avg_margin = sum(margins) / len(margins)
        print(f"   Average trade margin: {avg_margin:.1f}%  (across {len(margins)} priced products)")

    # ── Photo coverage report ────────────────────────────────────────────────
    print()
    if not os.path.isdir(photos_dir):
        print("📷 No 'Photos' folder found.")
        print("   Create one next to this script and add product images named")
        print("   exactly after the product (e.g.  AMOXYCLAV 625.jpg)")
    else:
        all_products   = [d["product"] for d in data if d["product"]]
        has_photo      = [p for p in all_products if sanitize_photo_name(p) in photo_map]
        missing_photos = [p for p in all_products if sanitize_photo_name(p) not in photo_map]

        print(f"📷 Photo coverage: {len(has_photo)} / {len(all_products)} products")

        if missing_photos and len(missing_photos) <= 20:
            print(f"\n   Missing photos ({len(missing_photos)} products):")
            for p in missing_photos:
                print(f"      → {p}.jpg")
        elif missing_photos:
            print(f"   {len(missing_photos)} products still need photos.")
        else:
            print("   ✓ All products have photos!")

        # Warn about unmatched photo files (typos in filenames)
        product_keys = {sanitize_photo_name(p) for p in all_products if p}
        unmatched = [v for k, v in photo_map.items() if k not in product_keys]
        if unmatched:
            print(f"\n   ⚠ Unmatched photo files (check for typos):")
            for f in unmatched:
                print(f"      ! Photos/{f}")

    print(f"\n→ Open SNT.html in your browser.\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert Master_File.xlsx → data.js for SNT catalogue")
    parser.add_argument("--file", default="Master_File.xlsx", help="Path to Excel file")
    parser.add_argument("--out",  default="data.js",          help="Output JS file name")
    args = parser.parse_args()
    convert(args.file, args.out)
